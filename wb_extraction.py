from selenium import webdriver  # Importing the Selenium webdriver module
from selenium.webdriver.common.keys import Keys  # Importing Keys from selenium.webdriver.common.keys for sending special keys like Enter, F1, etc.
from selenium.webdriver.common.by import By  # Importing By class from selenium.webdriver.common.by for locating elements
from selenium.webdriver.chrome.options import Options  # Importing Options from selenium.webdriver.chrome.options for configuring the Chrome webdriver options
from openpyxl import Workbook  # Importing Workbook from openpyxl module for working with Excel files
from selenium.webdriver.support import expected_conditions as EC  # Importing expected_conditions from selenium.webdriver.support to work with explicit waits
from selenium.webdriver.support.ui import WebDriverWait  # Importing WebDriverWait from selenium.webdriver.support.ui for implementing explicit waits
from selenium.webdriver import ActionChains  # Importing ActionChains from selenium.webdriver for performing actions like drag-and-drop, hover, etc.
import time  # Importing time module for time-related functions


# Setting up Chrome WebDriver
chrome_options = Options()  # Creating an instance of Chrome Options
chrome_driver_path = r"/home/bhargav/Downloads/chromedriver-linux64 (1)/chromedriver-linux64"  # Path to Chrome driver executable
driver = webdriver.Chrome(options=chrome_options)  # Initializing Chrome WebDriver with the specified options


#actual URL -- https://confluence.vc/all-investors/
# List of URLs to scrape
Link = ['https://investors.confluence.vc/advertising-marketing',
        'https://investors.confluence.vc/agriculture',
        'https://investors.confluence.vc/ai',
        'https://investors.confluence.vc/ar-vr',
        'https://investors.confluence.vc/arts',
        'https://investors.confluence.vc/beauty',
        'https://investors.confluence.vc/biotech',
        'https://investors.confluence.vc/cannabis',
        'https://investors.confluence.vc/chemicals',
        'https://investors.confluence.vc/cleantech',
        'https://investors.confluence.vc/cloud',
        'https://investors.confluence.vc/consumer',
        'https://investors.confluence.vc/content',
        'https://investors.confluence.vc/crypto-web3',
        'https://investors.confluence.vc/cybersecurity',
        'https://investors.confluence.vc/developer-tools',
        'https://investors.confluence.vc/e-commerce',
        'https://investors.confluence.vc/education',
        'https://investors.confluence.vc/electronics',
        'https://investors.confluence.vc/enterprise',
        'https://investors.confluence.vc/e-sports',
        'https://investors.confluence.vc/fashion',
        'https://investors.confluence.vc/fintech',
        'https://investors.confluence.vc/food-and-beverage',
        'https://investors.confluence.vc/food-services',
        'https://investors.confluence.vc/govtech',
        'https://investors.confluence.vc/hospitality',
        'https://investors.confluence.vc/hr',
        'https://investors.confluence.vc/insurtech',
        'https://investors.confluence.vc/legal',
        'https://investors.confluence.vc/logistics',
        'https://investors.confluence.vc/manufacturing',
        'https://investors.confluence.vc/marketplaces',
        'https://investors.confluence.vc/parenting',
        'https://investors.confluence.vc/proptech',
        'https://investors.confluence.vc/recruiting',
        'https://investors.confluence.vc/robotics',
        'https://investors.confluence.vc/sales',
        'https://investors.confluence.vc/services',
        'https://investors.confluence.vc/sex-tech',
        'https://investors.confluence.vc/social-impact',
        'https://investors.confluence.vc/space',
        'https://investors.confluence.vc/communications',
        'https://investors.confluence.vc/transportation',
        'https://investors.confluence.vc/health-and-wellness',
        'https://investors.confluence.vc/future-of-work',
        'https://investors.confluence.vc/internet-of-things',
        'https://investors.confluence.vc/quantum-computing',
        'https://investors.confluence.vc/travel',
        'https://investors.confluence.vc/sports']


for url in Link:  # Looping through each URL
    driver.get(url)  # Opening the webpage specified by the URL

    # Wait for the page to load here if necessary
    input("start")  # Pausing execution until user input (pressing Enter) to start scraping

    # Finding elements with the specified CSS selectors
    Person_name = driver.find_elements(By.CSS_SELECTOR, 'h1[data-id="_i7sulsj9u"]')  # Finding h1 elements with data-id attribute equal to "_i7sulsj9u"
    Person_role = driver.find_elements(By.CSS_SELECTOR, 'div[data-id="_i7sulsj9u12"]')  # Finding div elements with data-id attribute equal to "_i7sulsj9u12"
    Person_fund = driver.find_elements(By.CSS_SELECTOR, 'div[data-id="_kbtr0o72812"]')  # Finding div elements with data-id attribute equal to "_kbtr0o72812"
    Person_image = driver.find_elements(By.CSS_SELECTOR, 'div[data-type="staticImage"]')  # Finding div elements with data-type attribute equal to "staticImage"
    Person_logo = driver.find_elements(By.CSS_SELECTOR, 'div.sw-js-single-item-elements.w-100 div.sw-pre-image-container')  # Finding div elements with specific CSS selectors

    # Initializing empty lists to store extracted data
    Name_data, Role_data, Fund_data, Image_data, Logo_data = [], [], [], [], []

    # Extracting data from elements
    Name_data = [element.text for element in Person_name]  # Extracting text from h1 elements
    Role_data = [element.text for element in Person_role]  # Extracting text from div elements
    Fund_data = [element.text for element in Person_fund]  # Extracting text from div elements
    Image_data = [element.get_attribute('data-value') for element in Person_image]  # Finding div elements with data-type attribute equal to "staticImage"
     # Extracting 'data-value' attribute from div elements

    # Extracting image URLs from div elements and handling exceptions
    for element in Person_logo:
        try:
            x = element.find_element(By.CSS_SELECTOR, 'span[data-type="image"]')  # Finding span elements with data-type attribute equal to "image"
            Logo_data.append(x.get_attribute('data-value'))  # Extracting 'data-value' attribute and appending to the data list
        except Exception as e:
            print('logo not found')  # Handling case when logo is not found
            Logo_data.append(' ')  # Appending empty string if logo is not found

    # Creating a new Excel workbook
    wb = Workbook()  # Creating an instance of Workbook
    ws = wb.active  # Getting the active sheet of the workbook

    # Writing data to Excel sheet
    for i, (name_val, role_val, fund_val, image_val, logo_val) in enumerate(zip(Name_data, Role_data, Fund_data, Image_data, Logo_data), start=1):
        ws.cell(row=i, column=1, value=name_val)  # Writing data to first column
        ws.cell(row=i, column=2, value=role_val)  # Writing data to second column
        ws.cell(row=i, column=3, value=fund_val)  # Writing data to third column
        ws.cell(row=i, column=4, value=image_val)  # Writing data to fourth column
        ws.cell(row=i, column=5, value=logo_val)  # Writing data to fifth column

    # Saving the workbook with a filename based on the URL
    wb.save(f"{url.split('/')[-1]}.xlsx")  # Saving the workbook with a filename derived from the last part of the URL

    input("done")  # Pausing execution until user input (pressing Enter) after finishing scraping

# Quitting the driver
driver.quit()  # Closing the browser window and quitting the WebDriver
