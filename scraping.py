




'''
webdriver_path = "/home/bhargav/Downloads/chromedriver-linux64"

driver = webdriver.Chrome(webdriver_path)


url = 'https://investors.confluence.vc/advertising-marketing'
driver.get(url)

driver.implicitly_wait(10)


page_source = driver.page_source


driver.quit()
soup = BeautifulSoup(page_source, 'html.parser')

divs = soup.select_one('div.js-list-item.position-relative.overflow-hidden.p-0.flex-1.justify-content-center.box.sw-background-color-ffffff.sw-margin-bottom-none.sw-border-style-solid.sw-border-width-s.sw-border-color-e6e6e6.sw-border-radius-l.sw-box-shadow-none.sw-cursor-default.left')
for div in divs:
    print(div.get_text(strip=True))
'''




# url = "https://investors.confluence.vc/advertising-marketing"
# response = requests.get(url)

# soup = BeautifulSoup(response.content, 'lxml')
#print(soup.prettify())

# parent_div = soup.select_one('div.js-list-item.position-relative.overflow-hidden.p-0.flex-1.justify-content-center.box.sw-background-color-ffffff.sw-margin-bottom-none.sw-border-style-solid.sw-border-width-s.sw-border-color-e6e6e6.sw-border-radius-l.sw-box-shadow-none.sw-cursor-default.left')

# if parent_div:
#     parent_div_text = parent_div.get_text(separator='', strip=True)
#     print(parent_div_text)
# else:
#     print("not found")




# url = "https://investors.confluence.vc/advertising-marketing"
# response = requests.get(url)

# soup = BeautifulSoup(response.content, 'lxml')
# #print(soup.prettify())



# people = soup.select_one('div', class_="sw-js-list-container        sw-3-column          ")

# person1 = soup.select('div', class_="js-list-item position-relative overflow-hidden p-0 flex-1 justify-content-center box sw-background-color-ffffff sw-margin-bottom-none sw-border-style-solid sw-border-width-s sw-border-color-e6e6e6 sw-border-radius-l sw-box-shadow-none hover:sw-box-shadow-m sw-cursor-default  left ")
# #print(person1)

# person1_img = soup.select('div', class_="image text-center")# data-value="https://v5.airtableusercontent.com/v3/u/27/27/1711116000000/P3XH1fZLGSu_89cuvR3pWw/x5LMYPg7vLJqD1w8zJJWiETCzumzmeMOYMe-Q0hIB4z__cbcxE8_9JdHf7zF4fW4My0u5ab0uonjeJWQvNgD5uCLmyaSp6qvKCYR-mvTuaxXYpmDPXTWQG0mzT7HcaUqGBbvVvMto1tOLToopMtoVw/KoyZ6_8zLA4ieKYAwtGiF-98lEOWeiKj2s7RwIqQ4O4"     ## data-id="_htaykdl44" 
# #print(person1_img)

# person1_name = soup.select('div', class_="sw-pre-text-container",id = "_i7sulsj9u") #data-id="_i7sulsj9u"          
# #print(person1_name)

# person1_role = soup.select('div', class_="sw-pre-text-container" )#data-id="_i7sulsj9u12"
# #print(person1_role)

# person1_fund = soup.select('div', class_="sw-pre-text-container" )#data-id="_kbtr0o72812" 
# #print(person1_fund)






# # people = soup.select_one('div', class_="sw-js-list-container        sw-3-column          ")

# # person1 = soup.select('div', class_="js-list-item position-relative overflow-hidden p-0 flex-1 justify-content-center box sw-background-color-ffffff sw-margin-bottom-none sw-border-style-solid sw-border-width-s sw-border-color-e6e6e6 sw-border-radius-l sw-box-shadow-none hover:sw-box-shadow-m sw-cursor-default  left ")
# # #print(person1)

# # person1_img = soup.select('div', class_="image text-center")# data-value="https://v5.airtableusercontent.com/v3/u/27/27/1711116000000/P3XH1fZLGSu_89cuvR3pWw/x5LMYPg7vLJqD1w8zJJWiETCzumzmeMOYMe-Q0hIB4z__cbcxE8_9JdHf7zF4fW4My0u5ab0uonjeJWQvNgD5uCLmyaSp6qvKCYR-mvTuaxXYpmDPXTWQG0mzT7HcaUqGBbvVvMto1tOLToopMtoVw/KoyZ6_8zLA4ieKYAwtGiF-98lEOWeiKj2s7RwIqQ4O4"     ## data-id="_htaykdl44" 
# # #print(person1_img)

# # person1_name = soup.select('div', class_="sw-pre-text-container",id = "_i7sulsj9u") #data-id="_i7sulsj9u"          
# # #print(person1_name)

# # person1_role = soup.select('div', class_="sw-pre-text-container" )#data-id="_i7sulsj9u12"
# # #print(person1_role)

# # person1_fund = soup.select('div', class_="sw-pre-text-container" )#data-id="_kbtr0o72812" 
# # #print(person1_fund)






'''

person2 = soup.find_all('a', class_="card-click-overlay" )#href="https://www.linkedin.com/in/acacia-lawrie/" 

person2_img = soup.find('div', class_="image text-center") #data-value="https://v5.airtableusercontent.com/v3/u/27/27/1711116000000/vh0OeEqJdTWLgLKHlk9Sng/f6jMx314PyGMgLsIIZSF_PCWVRcCjm3SD5Ye70wp3XBLKvuyxLNzhjgYl9qZxLk4UnkSNsEKus5lhmzBfV_LHTDn1dmoA3Ei2prVisbzJO53A-H5fMhe_w58rEFNVy0ckhPll18J0YzsRVcEXF9_lA/ceGnjAU8UzTk8RAzclTPGOOfEjZuagc7Z1ZFInCeMI0"     ### data-id="_htaykdl44" 

#person2_txt = soup.find_all(<div class="text d-flex flex-column">      <div class="sw-js-single-item-elements w-100" data-recordid="recnKKo26w1cdnJVo">              <div class="sw-pre-text-container" style="height: 29.5938px;">        <h1 data-type="heading1" data-value="Acacia Lawrie" data-mappedto="Name" data-id="_i7sulsj9u" class="sw-font-size-l sw-text-color-ff7269 sw-font-family-open_sans sw-font-weight-semibold sw-text-align-center sw-letter-spacing-normal sw-padding-top-none sw-padding-bottom-6xs sw-padding-left-5xs sw-padding-right-6xs ">         Acacia Lawrie        </h1>       </div>                     <div class="sw-pre-text-container" style="height: 31px;">        <div data-type="text" data-value="Startup Investor" data-mappedto="Title" data-id="_i7sulsj9u12" class="sub-title sw-font-size-l sw-text-color-333333 sw-font-family-default sw-font-weight-normal sw-text-align-center sw-letter-spacing-normal sw-padding-top-none sw-padding-bottom-7xs sw-padding-left-5xs sw-padding-right-6xs ">         Startup Investor        </div>       </div>                      <div class="sw-pre-text-container" style="height: 39px;">        <div data-type="text" data-value="37 Angels" data-mappedto="Fund" data-id="_kbtr0o72812" class="sub-title sw-font-size-l sw-text-color-000000 sw-font-family-default sw-font-weight-semibold sw-text-align-center sw-letter-spacing-normal sw-padding-top-none sw-padding-bottom-5xs sw-padding-left-6xs sw-padding-right-6xs ">         37 Angels        </div>       </div>                                                                                               <div class="sw-pre-image-container" style="height: 79px;">                          <div class="sw-image-thumbnail-gallery prevent-style-img sw-width-9xs sw-height-9xs sw-background-size-cover sw-border-style-none sw-border-width-xs sw-border-color-000000 sw-border-radius-none sw-text-align-left sw-padding-top-none sw-padding-bottom-6xs sw-margin-left-6xs sw-margin-right-6xs sw-background-repeat-no-repeat sw-background-position-center">                     <a class="sw-img" href="https://v5.airtableusercontent.com/v3/u/27/27/1711116000000/8ef4uHXn9C-sPKjuUketng/ClzLwGTGaHM3wvB38Tzja_7dptmDu1paEvGATg5Mh4AUmYoKauhYmwyeRv6FctSwQL6oL6U8POzd8gxMLWZrksSTPGra-2yIT_bJF-RJ940ANpd46j_8nyFdkacXffKrOSP6qmqjYLGXpUSuOtQ86g/n4D0l1dNonfpsbSzITO6jyHbEj-ry3PP97JC5yazV8c">            <span data-type="image" data-value="https://v5.airtableusercontent.com/v3/u/27/27/1711116000000/8ef4uHXn9C-sPKjuUketng/ClzLwGTGaHM3wvB38Tzja_7dptmDu1paEvGATg5Mh4AUmYoKauhYmwyeRv6FctSwQL6oL6U8POzd8gxMLWZrksSTPGra-2yIT_bJF-RJ940ANpd46j_8nyFdkacXffKrOSP6qmqjYLGXpUSuOtQ86g/n4D0l1dNonfpsbSzITO6jyHbEj-ry3PP97JC5yazV8c" data-mappedto="Fund Logo" data-id="_qjcd2rnvx" class="sw-md-w-100 d-inline-block sw-width-9xs sw-height-9xs sw-background-size-cover sw-border-style-none sw-border-width-xs sw-border-color-000000 sw-border-radius-none sw-text-align-left sw-padding-top-none sw-padding-bottom-6xs sw-margin-left-6xs sw-margin-right-6xs sw-background-repeat-no-repeat sw-background-position-center" style="background-image: url(https://v5.airtableusercontent.com/v3/u/27/27/1711116000000/8ef4uHXn9C-sPKjuUketng/ClzLwGTGaHM3wvB38Tzja_7dptmDu1paEvGATg5Mh4AUmYoKauhYmwyeRv6FctSwQL6oL6U8POzd8gxMLWZrksSTPGra-2yIT_bJF-RJ940ANpd46j_8nyFdkacXffKrOSP6qmqjYLGXpUSuOtQ86g/n4D0l1dNonfpsbSzITO6jyHbEj-ry3PP97JC5yazV8c); padding: 0;">             </span>           </a>                   </div>               </div>                            <div class="sw-pre-need-container" style="height: 0px;"></div>      </div>     </div>)

person2_name = soup.find('div', class_="sw-pre-text-container") #data-id="_i7sulsj9u"

person2_role = soup.find('div', class_="sw-pre-text-container") #data-id="_i7sulsj9u12" 

person2_fund = soup.find('div', class_="sw-pre-text-container") #data-id="_kbtr0o72812" 










person3 = soup.find_all('a', class_="card-click-overlay") #href="https://www.linkedin.com/in/adam-lebovitz-400b5069/"

person3_img = soup.find('div', class_="image text-center") #data-value="https://v5.airtableusercontent.com/v3/u/27/27/1711116000000/ARae8JJcC7zwmIJbiZb36A/mOlM5Gl-ebZ8xIa9xRz0E53Vl1YHFFfwUmUcEoXMvOpI7cyYuOdQKRKnAlEKXb9Tv6meKlYYEpNUkisy84OV_hebyxtjc6cG7PmWQbKJzmlZsndaLjT56y6FBy2zcjasOk-vhN5ahlK5EDmrCFwMog/KNJLuQojk14k6tJ5S4s1QcYr_5sxm41uK91MLKDuXQ8"
#data-id="_htaykdl44" 
                        
person3_name = soup.find('div', class_="sw-pre-text-container") #data-id="_i7sulsj9u" 

person3_role = soup.find('div', class_="sw-pre-text-container") #data-id="_i7sulsj9u12"

person3_fund = soup.find('div', class_="sw-pre-text-container") #data-id="_kbtr0o72812" 








person4 = soup.find_all('div', recordid="rec7Q2KYEF24yL0Mu" ,class_="js-list-item position-relative overflow-hidden p-0 flex-1 justify-content-center box sw-background-color-ffffff sw-margin-bottom-none sw-border-style-solid sw-border-width-s sw-border-color-e6e6e6 sw-border-radius-l sw-box-shadow-none hover:sw-box-shadow-m sw-cursor-default  left ")
# href="https://www.linkedin.com/in/adam-ned-0569ab180/"                     

person4_img = soup.find('div', class_="image text-center") #data-value="https://v5.airtableusercontent.com/v3/u/27/27/1711116000000/aRK6ZVKBES0zyjHZ3LffNg/wQDoAylks7JSycOYFRUpW-id-LCYVSQPYOPoMbx5n8rS6LuzzO0MwWpq8z8Q15e3nmTJhcsxRiAdR9IvKK6VcHJIcERIIc6mf_fqgR3_8cnffLr_2Uek-ceZGqtm_PYa9gL_jR4nh1OOMmhLPfi9OA/GsROmTvGza2WURVyufHZc5MNzgLcabwU10DNBt6E8hI"
#background-image: url(https://v5.airtableusercontent.com/v3/u/27/27/1711116000000/aRK6ZVKBES0zyjHZ3LffNg/wQDoAylks7JSycOYFRUpW-id-LCYVSQPYOPoMbx5n8rS6LuzzO0MwWpq8z8Q15e3nmTJhcsxRiAdR9IvKK6VcHJIcERIIc6mf_fqgR3_8cnffLr_2Uek-ceZGqtm_PYa9gL_jR4nh1OOMmhLPfi9OA/GsROmTvGza2WURVyufHZc5MNzgLcabwU10DNBt6E8hI)
#data-id="_htaykdl44"

person4_name = soup.find('div', class_="sw-pre-text-container") #data-id="_i7sulsj9u" 

person4_role = soup.find('div', class_="sw-pre-text-container") #data-id="_i7sulsj9u12" 

person4_fund = soup.find('div', class_="sw-pre-text-container") #data-id="_kbtr0o72812" 







person5 = soup.find_all('a', class_="card-click-overlay") #href="https://www.linkedin.com/in/adam-piasecki-/" 

person5_img = soup.find('div', class_="image text-center") #data-id="_htaykdl44" 
#data-value="https://v5.airtableusercontent.com/v3/u/27/27/1711116000000/EmTeXNGRVOglkqa4uDFWDw/E67c-ox5pr1NGEpmKADisPnnulcc8AexOweT4cSNfp2da82zEhYF6sQ0WI93Cvaz1LEk38DVl761o8gV_gCWf988idjKyBEioll3jRpCBwJX3fLY3hVFiXOHAzBkL0E3_FSQnG2yQiZa3yrwpyi3B6rIEoUNHqUol7y6tDuqjT4/bLQhQGDb3b_c0GU8wBC0_Pl9GvXZD4Zn0ZF7imk9pmo" 

person5_name = soup.find('div', class_="sw-pre-text-container") #data-id="_i7sulsj9u" 

person5_role = soup.find('div', class_="sw-pre-text-container") #data-id="_i7sulsj9u12" 

person5_fund = soup.find('div', class_="sw-pre-text-container") #data-id="_kbtr0o72812" 








###2ndtask  
element_text = ["person_1:Abigail Malin", "person_2:Acacia Lawrie", "person_3:Adam Lebovitz", "person_4:Adam Ned", "person_5:Adam Piasecki" ]
element_link = ["link1:https://www.linkedin.com/in/abigailmalin/", "link2:https://www.linkedin.com/in/acacia-lawrie/", 
                "link3:https://www.linkedin.com/in/adam-lebovitz-400b5069/", "link4:https://www.linkedin.com/in/adam-ned-0569ab180/",
                "link5:href=https://www.linkedin.com/in/adam-piasecki-/"]

ele_dict = {}#empty dict for storing key-value pairs
 
for text, link in zip(element_text, element_link):
    ele_dict[text] = link

print(ele_dict)







create_table = pd.DataFrame{'persons': ['person 1', 'person 2', 'person 3', 'person 4', 'person 5'],
                            'Name': ['Abigail Malin', 'Acacia Lawrie', '']}'''


















# from bs4 import BeautifulSoup
# import requests
# import pandas as pd


# try:
#     web_url = "https://confluence.vc/all-investors/"
#     content = requests.get(web_url)
#     #print(content)#.raise_for_status())

#     soup = BeautifulSoup(web_url,'lxml')
#     #print(soup)


#     persons = soup.find_all('div', class_="sw-js-list-container.sw-3-column")
#     print(persons)




# except Exception as e:
#     print(e)






















from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains
import pandas as pd


# Set up Chrome WebDriver
chrome_options = Options()
chrome_driver_path = r"/home/bhargav/Downloads/chromedriver-linux64 (1)/chromedriver-linux64"  # Path to chromedriver
service = Service(chrome_driver_path)
driver = webdriver.Chrome(options=chrome_options)

time.sleep(10)


# URL of the investor profile
url = "https://signal.nfx.com/investors/eric-chen"

# Initialize Selenium WebDriver (you may need to specify the path to your webdriver executable)

driver.get(url)

# # Scraping investor details
# name = driver.find_element(By.CSS_SELECTOR, 'h1.f3.f1-ns.mv1')#.text()  #).strip()
# #link = driver.find_element(By.CSS_SELECTOR, f'a[class="mr3"]').text()#.strip()
# address = driver.find_element(By.CSS_SELECTOR, 'span.ml1') #.text() #.strip()
# current_investing_position = driver.find_element(By.CSS_SELECTOR, 'div.line-separated-row.row')#.text()#.strip()
# invest_range = driver.find_element(By.CSS_SELECTOR, 'div.line-separated-row.row')#.text()
# sweet_spot = driver.find_element(By.CSS_SELECTOR, 'div.line-separated-row.row')#.text()
# investments_on_record = driver.find_element(By.CSS_SELECTOR, 'div.line-separated-row.row') #.text()
# current_fund_size = driver.find_element(By.CSS_SELECTOR, 'div.line-separated-row.row')#.text()
# sector_and_stage_rankings = driver.find_elements(By.CSS_SELECTOR, 'a.vc-list-chip')#.text()



# Wait for the element to be present
wait = WebDriverWait(driver, 10)
name_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h1.f3.f1-ns.mv1')))
#link--------------------------------------
address_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'span.ml1')))
current_investing_position_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.line-separated-row.row')))
invest_range_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.line-separated-row.row')))
sweet_spot_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.line-separated-row.row')))
investments_on_record_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.line-separated-row.row')))
current_fund_size_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.line-separated-row.row')))
sector_and_stage_rankings_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a.vc-list-chip')))



name = name_element.text.strip()
address = address_element.text.strip()
current_investing_position = current_investing_position_element.text.strip()
invest_range = invest_range_element.text.strip()
sweet_spot = sweet_spot_element.text.strip()
investments_on_record = investments_on_record_element.text.strip()
current_fund_size = current_fund_size_element.text.strip()
sector_and_stage_rankings = sector_and_stage_rankings_element.text.strip()

print(current_investing_position)
# Close the WebDriver

# Create a dictionary with scraped data
investor_data = {
    'Name': name,
    #'Link': link,
    'Address': address,
    'Current Investing Position': current_investing_position,
    'Investment Range': invest_range,
    'Sweet Spot': sweet_spot,
    'Investments On Record': investments_on_record,
    'Current Fund Size': current_fund_size,
    'Sector and Stage Rankings': sector_and_stage_rankings
}

# # Convert scraped data into a DataFrame
# df = pd.DataFrame([investor_data])

# # Write DataFrame to Excel file
# df.to_excel("investor_data.xlsx", index=False)
 
# Create a new Excel workbook
wb = Workbook() 
ws = wb.active

# Write data to Excel sheet
for i, (value1, value2, value3, value4, value5, value6, value7, value8) in enumerate(zip(name, address, current_investing_position, invest_range, sweet_spot, investments_on_record, current_fund_size, sector_and_stage_rankings), start=1):
    ws.cell(row=i, column=1, value=value1)
    ws.cell(row=i, column=2, value=value2)
    ws.cell(row=i, column=3, value=value3)
    ws.cell(row=i, column=4, value=value4)
    ws.cell(row=i, column=5, value=value5) 
    ws.cell(row=i, column=6, value=value6) 
    ws.cell(row=i, column=7, value=value7) 
    ws.cell(row=i, column=8, value=value8) 


# Save the workbook
wb.save("signalnfx.xlsx")

driver.quit()





'''
  

# Extract data from elements
organisations_data = []
for element in zip(name_f, brief_f, description_f, address_f, links_text, links_f):
    name = element[0].text
    brief = element[1].text
    description = element[2].text
    address = element[3].text
    links_text = element[4].text
    links = element[5].get_attribute('href')
    organisations_data.append({"Name": name, "Brief": brief, "Description": description, "IN A SNAPSHOT": address, "Link": links_text, "Links": links})
    print(organisations_data)



# Create a new Excel workbook
wb = Workbook()
ws = wb.active

# Write data to Excel sheet
ws.append(["Name", "Brief", "Description", "IN A SNAPSHOT", "Link", "Links"])
for organisation in organisations_data:
    ws.append([organisation["Name"], organisation["Brief"], organisation["Description"], organisation["IN A SNAPSHOT"], organisation["Link"], organisation["Links"]])

# Save the workbook
wb.save(f"{url.split('/')[-1]}.xlsx")
input("done")




'''