from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains


# # Set up Chrome WebDriver
# chrome_options = Options()
# service = Service('home//bhargav//Downloads//chromedriver-linux64 (1)//chromedriver-linux64.exe')  # Path to chromedriver
# driver = webdriver.Chrome(service=service, options=chrome_options)


# Set up Chrome WebDriver
chrome_options = Options()
# Update the path to your ChromeDriver for Linux
# chrome_options.add_argument("--headless")  # Add any other options you need

chrome_driver_path = r"/home/bhargav/Downloads/chromedriver-linux64 (1)/chromedriver-linux64"
driver = webdriver.Chrome(options=chrome_options)



# Website URL
url = "https://investors.confluence.vc/advertising-marketing"


# Open the webpage
driver.get(url)

# Wait for the page to load
time.sleep(5)

# # Enter password if there's a password prompt
# password_input = driver.find_element(By.NAME, "password")
# password_input.send_keys(password)
# password_input.send_keys(Keys.RETURN)

# Wait for the page to load after entering password
time.sleep(31)



# Find the elements with the specified column ids
column_elements1 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldS3CcrHFsPD5xMK"]')
column_elements2 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldLyx44QWmxoL04y"]')
column_elements3 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldhllZJvoeI6rM5w"]')
column_elements4 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldqaGgbrIRNOmBCq"]')
column_elements5 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldzBFxNGMUaTx6Zm"]')
column_elements6 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldAlHpXZtEPBQ4lx"]')
column_elements7 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldbaUYvsHjxcDDbn"]')
column_elements8 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldpmw5AKCI1GkUKd"]')
column_elements9 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldkNzI7X1v4MwpLk"]')
column_elements10 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldGEGwb6eHM3xIqZ"]')
column_elements11 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldoUH1KXovlDOISZ"]')
column_elements12 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fld9jCaUDyHvkakcw"]')
column_elements13 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldOaR8QcHIrRm6jp"]')
column_elements14 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldn8uNumLCPfz0YM"]')
column_elements15 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldt99ErlOzvPqEX4"]')
column_elements16 = driver.find_elements(By.CSS_SELECTOR, f'div[data-columnid="fldMKVOobuG7Z3YuA"]')

# Extract data from elements
data1 = [element.text for element in column_elements1]
data2 = [element.text for element in column_elements2]
data3 = [element.text for element in column_elements3]
data4 = [element.text for element in column_elements4]
data5 = [element.text for element in column_elements5]
data6 = [element.text for element in column_elements6]
data7 = [element.text for element in column_elements7]
data8 = [element.text for element in column_elements8]
data9 = [element.text for element in column_elements9]
data10 = [element.text for element in column_elements10]
data11 = [element.text for element in column_elements11]
data12 = [element.text for element in column_elements12]
data13 = [element.text for element in column_elements13]
data14 = [element.text for element in column_elements14]
data15 = [element.text for element in column_elements15]
data16 = [element.text for element in column_elements16]

data3_script = '''
var elements = document.querySelectorAll('div[data-columnid="fldhllZJvoeI6rM5w"]');
complete_content = []

// Loop through the elements

for (var i = 0; i < elements.length; i++) {
    var element = elements[i];
    var element_list = element.querySelectorAll('*>span')
    var span_content = []
    for (var j = 0; j < element_list.length; j++){
        span_content.push(element_list[j].textContent);
}
    complete_content.push(span_content)
}
return complete_content
'''
data3 = driver.execute_script(f'  {data3_script}')

data5_script = '''
var elements = document.querySelectorAll('div[data-columnid="fldzBFxNGMUaTx6Zm"]');
complete_content = []

// Loop through the elements

for (var i = 0; i < elements.length; i++) {
    var element = elements[i];
    var element_list = element.querySelectorAll('*>span')
    var span_content = []
    for (var j = 0; j < element_list.length; j++){
        span_content.push(element_list[j].textContent);
}
    complete_content.push(span_content)
}
return complete_content
'''
data5 = driver.execute_script(f'  {data5_script}')

#data5_string=str(data5)



data8_script = '''
var elements = document.querySelectorAll('div[data-columnid="fldpmw5AKCI1GkUKd"]');
complete_content = []

// Loop through the elements

for (var i = 0; i < elements.length; i++) {
    var element = elements[i];
    var element_list = element.querySelectorAll('*>div')
    if (elements.lenght==0){
         continue;
    }
    var span_content = []
    for (var j = 0; j < element_list.length; j++){
        span_content.push(element_list[j].textContent);
}
    complete_content.push(span_content)
}
return complete_content'''

data8 = driver.execute_script(f' {data8_script}')

data9_script = '''
var elements = document.querySelectorAll('div[data-columnid="fldkNzI7X1v4MwpLk"]');
complete_content = []

// Loop through the elements

for (var i = 0; i < elements.length; i++) {
    var element = elements[i];
    var element_list = element.querySelectorAll('*>span')
    var span_content = []
    for (var j = 0; j < element_list.length; j++){
        span_content.push(element_list[j].textContent);
}
    complete_content.push(span_content)
}
return complete_content
'''
data9 = driver.execute_script(f'  {data9_script}')
#data8_string=str(data8)

# Find all scroll bar elements
scroll_bar = WebDriverWait(driver, 500).until(
    EC.visibility_of_element_located((By.CSS_SELECTOR, ".antiscroll-scrollbar.antiscroll-scrollbar-vertical.antiscroll-scrollbar-shown"))
)


# Create an ActionChains object
actions = ActionChains(driver)

# Click and hold the scroll bar element
actions.click_and_hold(scroll_bar).perform()
actions.move_by_offset(0,0).perform()
actions.release().perform()
time.sleep(5)


# Create a new Excel workbook
wb = Workbook() 
ws = wb.active
# for i in range(1, 17):
#     variable_name = f"data{i}"
#     variable_value = globals()[variable_name]
#     print(variable_name, variable_value)



# Write data to Excel sheet
for i, (value1, value2, value3, value4, value5, value6,value7,value8,value9,value10,value11,value12,value13,value14,value15,value16) in enumerate(zip(data1, data2, data3, data4, data5,data6,data7,data8,data9,data10,data11,data12,data13, data14,data15,data16), start=1):
    ws.cell(row=i, column=1, value=value1)
    ws.cell(row=i, column=2, value=value2)
    value3_sub = '; '.join(value3)
    ws.cell(row=i, column=3, value=value3_sub)
    ws.cell(row=i, column=4, value=value4)
    value5_sub = '; '.join(value5)
    ws.cell(row=i, column=5, value=value5_sub)
    ws.cell(row=i, column=6, value=value6)
    ws.cell(row=i, column=7, value=value7)
    value8_sub= '; '.join(value8)
    ws.cell(row=i, column=8, value=value8_sub)
    value9_sub = '; '.join(value9)
    ws.cell(row=i, column=9, value=value9_sub)
    ws.cell(row=i, column=10, value=value10)
    ws.cell(row=i, column=11, value=value11)
    ws.cell(row=i, column=12, value=value12)
    ws.cell(row=i, column=14, value=value14)
    ws.cell(row=i, column=15, value=value15)
    ws.cell(row=i, column=16, value=value16)


# Save the workbook
wb.save("datatest.xlsx")

# Quit the driver
driver.quit()