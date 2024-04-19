'''

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains

# Set up Chrome WebDriver
chrome_options = Options()

chrome_driver_path = r"/home/bhargav/Downloads/chromedriver-linux64 (1)/chromedriver-linux64"
driver = webdriver.Chrome(options=chrome_options)

# Website URL
url = "https://investors.confluence.vc/agriculture-tech"

input("start")

# Open the webpage
driver.get(url)
# Wait for the page to load

input("Enter to start Extracting")

# Find the elements with the specified column ids
column_elements1 = driver.find_elements(By.CSS_SELECTOR, f'h1[data-id="_i7sulsj9u"]')
column_elements2 = driver.find_elements(By.CSS_SELECTOR, f'div[data-id="_i7sulsj9u12"]')
column_elements3 = driver.find_elements(By.CSS_SELECTOR, f'div[data-id="_kbtr0o72812"]')
column_elements4 = driver.find_elements(By.CSS_SELECTOR, f'div[data-type="staticImage"]')
column_elements5 = driver.find_elements(By.CSS_SELECTOR, f'//div[@class="sw-js-single-item-elements w-100"]//div[@class="sw-pre-image-container"]//span')
#f'span[data-type="Image"])


# Extract data from elements
data1 = [element.text for element in column_elements1]
data2 = [element.text for element in column_elements2]
data3 = [element.text for element in column_elements3]
data4 = [element.get_attribute('data-value') for element in column_elements4]
data5 = [element.get_attribute('data-value') for element in column_elements5]


print(data2)

# Create a new Excel workbook
wb = Workbook() 
ws = wb.active

# Write data to Excel sheet
for i, (value1, value2, value3, value4, value5) in enumerate(zip(data1, data2, data3, data4, data5), start=1):
    ws.cell(row=i, column=1, value=value1)
    ws.cell(row=i, column=2, value=value2)
    ws.cell(row=i, column=3, value=value3)
    ws.cell(row=i, column=4, value=value4)
    ws.cell(row=i, column=5, value=value5) 

# Save the workbook
wb.save("Agriculture.xlsx")

# Quit the driver
driver.quit()

'''









from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import time

# Set up Chrome WebDriver
chrome_options = Options()
chrome_driver_path = r"/home/bhargav/Downloads/chromedriver-linux64 (1)/chromedriver-linux64"  # Path to chromedriver
service = Service(chrome_driver_path)
driver = webdriver.Chrome(options=chrome_options)

# Website URL
url = "https://climatescape.org/organizations/offsetra"

time.sleep(10)
# Open the webpage
driver.get(url)

# Define locators for elements
name_locator = (By.CSS_SELECTOR, '.flex-grow.text-xl.font-semibold')  # Locator for name
brief_locator = (By.CSS_SELECTOR, 'p')  # Locator for brief
description_locator = (By.XPATH, "(//div[@class='my-6'])[last()]")  # Locator for description
new_column_locator = (By.CSS_SELECTOR, '.flex.flex-col.mb-8')  # Locator for new column
additional_column_locator = (By.CSS_SELECTOR, '.svg-inline--fa fa-external-link-alt fa-w-16')  # Locator for additional column

# Wait for the elements to be present
wait = WebDriverWait(driver, 10)
name_element = wait.until(EC.visibility_of_element_located(name_locator))
brief_element = wait.until(EC.visibility_of_element_located(brief_locator))
description_element = wait.until(EC.visibility_of_element_located(description_locator))
new_column_element = wait.until(EC.visibility_of_element_located(new_column_locator))
additional_column_element = wait.until(EC.visibility_of_element_located(additional_column_locator))

# Extract text from elements
name_text = name_element.text
brief_text = brief_element.text
description_text = description_element.text
new_column_text = new_column_element.text
additional_column_text = additional_column_element.text

# Create a new Excel workbook
wb = Workbook()
ws = wb.active

# Write data to Excel sheet
try:
    ws.cell(row=1, column=1, value="Name")
    ws.cell(row=1, column=2, value="Brief")
    ws.cell(row=1, column=3, value="Description")
    ws.cell(row=1, column=4, value="New Column")
    ws.cell(row=1, column=5, value="Additional Column")

    ws.cell(row=2, column=1, value=name_text)
    ws.cell(row=2, column=2, value=brief_text)
    ws.cell(row=2, column=3, value=description_text)
    ws.cell(row=2, column=4, value=new_column_text)
    ws.cell(row=2, column=5, value=additional_column_text)

    # Save the workbook
    wb.save("data.xlsx")
    print("Data saved successfully!")
except Exception as e:
    print("Error occurred while saving data to Excel:", e)

# Quit the driver
driver.quit()